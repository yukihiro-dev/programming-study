import openpyxl										# Pythonに「Excelを操作する専門知識」を読み込ませるという命令

file_name = "出欠記録.xlsx"								# 変数に活用するExcelのファイル名を文字列で入れる。

try:											# 【例外処理】ファイル未検出などの実行時エラーを補足するためのブロック
    target_month = input("集計したい月を数字で入力してください（例: 2）: ")		# 標準出力:ユーザー指定の値を文字列として取得
    wb = openpyxl.load_workbook(file_name)						# openpyxlのメソッドを利用し、Excelファイルをメモリ上のオブジェクトにロード
    ws = wb.active 									# ワークブックの「現在のアクティブシート」プロパティを操作対象として保持
    
    daily_data = {}      # 指定した月の詳細用
    monthly_count = 0    # 指定した月の延べ人数
    monthly_total = 0    # 指定した月の合計金額
    yearly_total = 0     # 2026年すべての合計金額
    
    # 1. 全ての列をループして集計
    for j in range(3, ws.max_column + 1):
        date_val = ws.cell(row=1, column=j).value
        if date_val is None: continue
        
        # 月の判定（Excelの日付形式から抜き出し）
        try:
            current_month = str(date_val.month) if hasattr(date_val, 'month') else str(date_val).split('/')[0]
        except:
            continue

        # その列の「日計」と「人数」を計算
        day_sum = 0
        day_count = 0
        for i in range(2, ws.max_row + 1):
            fee = ws.cell(row=i, column=2).value
            attendance = ws.cell(row=i, column=j).value
            if str(attendance).strip() in ["◯", "○", "〇", "o", "O", "0"]:
                day_sum += int(fee) if fee else 0
                day_count += 1
        
        # 年間累計には全ての月を加算
        yearly_total += day_sum
        
        # 指定した月なら詳細に記録
        if current_month == target_month:
            if hasattr(date_val, 'strftime'):
                # %mは月(01)、%dは日(09)のように、一桁でも0をつけてくれます
                display_date = date_val.strftime('%m月%d日')
            else:
                display_date = str(date_val)
            
            daily_data[j] = {"date": display_date, "count": day_count, "amount": day_sum}
            monthly_count += day_count
            monthly_total += day_sum
  # 3. 結果の表示
    print(f"\n--- 【{target_month}月分】 集計レポート ---")
    if not daily_data:
        print(f"※{target_month}月の予定データは見つかりませんでした。")
    else:
        # 指定月の詳細を表示
        for j in daily_data:
            d = daily_data[j]
            if d['count'] > 0:
                print(f"{d['date']} 参加人数{d['count']}人 参加費合計{d['amount']}円")
        
        print(f"\n{target_month}月 参加人数{monthly_count}人 参加費総額{monthly_total}円")

        # 年間総額は、どの月を集計しても最後に必ず表示します
    print(f"2026年 参加費総額{yearly_total}円")

except Exception as e:
    print(f"エラー: {e}")